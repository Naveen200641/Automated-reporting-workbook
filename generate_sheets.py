import os
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font

wb = Workbook()

ws = wb.active
ws.title = "Sales_Data"
headers = ["Date", "Product", "Quantity", "Price", "Revenue"]
for c, h in enumerate(headers, 1):
    ws.cell(1, c, h).font = Font(bold=True)

data = [
    ["01-01-2025", "Laptop", 2, 50000],
    ["02-01-2025", "Mouse", 5, 500],
    ["03-01-2025", "Keyboard", 3, 1500],
    ["04-01-2025", "Monitor", 1, 12000],
    ["05-01-2025", "Laptop", 1, 50000],
    ["06-01-2025", "Mouse", 10, 500],
    ["07-01-2025", "Keyboard", 4, 1500],
    ["08-01-2025", "Monitor", 2, 12000],
]
for r, row in enumerate(data, 2):
    for c, v in enumerate(row, 1):
        ws.cell(r, c, v)
    ws.cell(r, 5, f"=C{r}*D{r}")

dash = wb.create_sheet("Dashboard")
dash["A1"] = "Automated Reporting Workbook"
dash["A1"].font = Font(bold=True, size=14)
dash["A3"] = "Total Revenue"
dash["B3"] = "=SUM(Sales_Data!E2:E100)"
dash["A4"] = "Total Quantity"
dash["B4"] = "=SUM(Sales_Data!C2:C100)"

chart = BarChart()
data_ref = Reference(ws, min_col=5, min_row=1, max_row=9)
cats = Reference(ws, min_col=2, min_row=2, max_row=9)
chart.add_data(data_ref, titles_from_data=True)
chart.set_categories(cats)
dash.add_chart(chart, "D2")

workspace_dir = r"c:\Users\navee\Desktop\Automated"
xlsx_path = os.path.join(workspace_dir, "Automated_Reporting_Workbook.xlsx")
wb.save(xlsx_path)

readme_path = os.path.join(workspace_dir, "README.md")
with open(readme_path, "w", encoding="utf-8") as f:
    f.write("# Automated Reporting Workbook\n\nExcel-based automated sales reporting dashboard with formulas and charts.")

print(f'{{"xlsx": "{xlsx_path.replace(os.sep, "/")}", "readme": "{readme_path.replace(os.sep, "/")}"}}')
